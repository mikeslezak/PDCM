#!/usr/bin/env python3
"""
S32K358 HDQFP-172 Pin Allocator for PDCM
=========================================
Reads the NXP S32K358 IOMUX spreadsheet, performs complete pin allocation,
and generates both a C header file and a text allocation report.

Usage: /c/Python314/python s32k358_pin_allocator.py
"""

import openpyxl
import re
import os
from collections import defaultdict
from datetime import datetime

IOMUX_PATH = "C:/Users/mikes/Downloads/S32K358_IOMUX.xlsx"
HEADER_OUT = "C:/Users/mikes/PDCM/firmware/shared/PDCMConfig_S32K358.h"
REPORT_OUT = "C:/Users/mikes/PDCM/hardware/schematic/pin_allocation_report.txt"


# =============================================================================
# Step 1: Parse the IOMUX spreadsheet
# =============================================================================

def parse_iomux():
    """Parse all relevant sheets from the IOMUX spreadsheet."""
    wb = openpyxl.load_workbook(IOMUX_PATH, data_only=True)

    # --- Parse Pinout sheet (pad numbers, pin names, alt functions) ---
    ws = wb['S32K358_Pinout']
    pinout = {}  # pin_name -> {pad172, alts: {ALTn: func}, direct_signals, composite}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        pin_name = row[2].value
        if not pin_name or pin_name == 'Pin Name':
            continue
        pad172 = row[0].value
        alts = {}
        for i in range(4, 21):
            val = row[i].value
            if val:
                header = ws.cell(row=2, column=i+1).value
                if header:
                    alts[header] = val
        # Direct signals are in col 21 (index 20)
        direct = row[20].value if len(row) > 20 else None
        composite = row[45].value if len(row) > 45 else None
        pad_type = row[46].value if len(row) > 46 else None
        pinout[pin_name] = {
            'pad172': pad172,
            'alts': alts,
            'direct': direct,
            'composite': str(composite) if composite else '',
            'pad_type': str(pad_type) if pad_type else ''
        }

    # --- Parse IO Signal Table (GPIO defaults, pad types, SSS values) ---
    ws2 = wb['S32K358_IO Signal Table']
    gpio_pins = {}  # pin_name -> {pad172, pad_type, gpio_num}
    last_port = None
    all_io_rows = []
    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
        port = row[0]
        if port:
            last_port = port
        sss = row[2]
        func = row[3]
        pad_type = row[7]
        pad172 = row[8]
        if sss == '0000_0000' and last_port:
            gpio_pins[last_port] = {
                'pad172': pad172,
                'pad_type': str(pad_type) if pad_type else '',
                'func': str(func) if func else ''
            }
        all_io_rows.append({
            'port': last_port,
            'sss': sss,
            'func': func,
            'pad172': pad172
        })

    # --- Parse PeripheralSummaries (ADC channel mapping) ---
    ws3 = wb['S32K358_PeripheralSummaries']
    adc_map = {}  # pad172 -> [(adc_instance, channel_name, channel_type)]
    current_adc = None
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
        pad172, bga, func, direction = row[0], row[1], row[2], row[3] if len(row) > 3 else None
        if func and str(func) in ('ADC0', 'ADC1', 'ADC2'):
            current_adc = str(func)
            continue
        if func and current_adc and str(func).startswith(current_adc):
            if direction == 'I' and pad172 and str(pad172) != '-':
                # Parse function name: ADC0_P4, ADC0_S17, ADC0_X[3], etc.
                func_str = str(func).strip()
                # Extract ADC instance number
                m = re.match(r'ADC(\d+)_([PSX])(\[?\d+\]?)', func_str)
                if m:
                    inst = int(m.group(1))
                    ch_type = m.group(2)  # P=precision, S=standard, X=external
                    ch_num_str = m.group(3).strip('[]')
                    ch_num = int(ch_num_str)
                    pad_int = int(pad172)
                    if pad_int not in adc_map:
                        adc_map[pad_int] = []
                    adc_map[pad_int].append((inst, ch_type, ch_num, func_str))
        elif current_adc and func and not str(func).startswith('ADC'):
            if str(pad172).startswith('S32K358'):
                current_adc = None  # New section header

    # --- Parse eMIOS from PeripheralSummaries ---
    emios_map = {}  # pad172 -> [(instance, channel, suffix)]
    current_emios = None
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
        pad172, bga, func, direction = row[0], row[1], row[2], row[3] if len(row) > 3 else None
        if func and str(func) in ('eMIOS_0', 'eMIOS_1', 'eMIOS_2'):
            current_emios = str(func)
            continue
        if func and current_emios and str(func).startswith(current_emios):
            func_str = str(func).strip()
            # May have multiple functions separated by newline
            for f in func_str.split('\n'):
                f = f.strip()
                m = re.match(r'eMIOS_(\d+)_CH\[(\d+)\]_([XHGY])', f)
                if m and pad172 and str(pad172) != '-' and pad172 is not None:
                    inst = int(m.group(1))
                    ch = int(m.group(2))
                    suffix = m.group(3)
                    pad_int = int(pad172)
                    if pad_int not in emios_map:
                        emios_map[pad_int] = []
                    emios_map[pad_int].append((inst, ch, suffix))
        elif current_emios and func and not str(func).startswith('eMIOS'):
            if str(pad172).startswith('S32K358'):
                current_emios = None

    wb.close()
    return pinout, gpio_pins, adc_map, emios_map


# =============================================================================
# Step 2: Build pin database with capabilities
# =============================================================================

def build_pin_db(pinout, gpio_pins, adc_map, emios_map):
    """Build a unified pin database with all capabilities."""
    pins = {}  # pin_name -> full info dict

    for pin_name, info in gpio_pins.items():
        pad = info['pad172']
        if pad == '-' or pad is None:
            continue
        try:
            pad_int = int(pad)
        except (ValueError, TypeError):
            continue

        # Parse port letter and number from pin name
        m = re.match(r'PT([A-H])(\d+)', pin_name)
        if not m:
            continue
        port_letter = m.group(1)
        port_num = int(m.group(2))

        # Determine if this is GPI-only (input-only, cannot be output)
        # Must distinguish "GPI" pad type from "GPIO-*" pad types
        pad_type_str = info.get('pad_type', '')
        func_str = info.get('func', '')
        is_gpi = (pad_type_str == 'GPI' or
                  func_str.startswith('GPI[') or
                  (pad_type_str.startswith('GPI') and not pad_type_str.startswith('GPIO')))
        # Check pinout sheet too
        if pin_name in pinout:
            po = pinout[pin_name]
            if po.get('alts', {}).get('ALT0', '').startswith('GPI['):
                is_gpi = True

        # Get ADC capabilities
        adc_channels = adc_map.get(pad_int, [])

        # Get eMIOS capabilities
        emios_channels = emios_map.get(pad_int, [])

        # Get alt function info from pinout sheet
        alt_funcs = {}
        composite = ''
        if pin_name in pinout:
            alt_funcs = pinout[pin_name].get('alts', {})
            composite = pinout[pin_name].get('composite', '')

        pins[pin_name] = {
            'pad': pad_int,
            'port': port_letter,
            'port_num': port_num,
            'pad_type': info['pad_type'],
            'is_gpi': is_gpi,
            'has_adc': len(adc_channels) > 0,
            'adc': adc_channels,
            'emios': emios_channels,
            'alt_funcs': alt_funcs,
            'composite': composite,
            'assigned': None,  # Will be filled during allocation
        }

    return pins


# =============================================================================
# Step 3: Perform pin allocation
# =============================================================================

def find_emios1_alt(pin_info, target_ch):
    """Find which ALT mode gives eMIOS_1_CH[target_ch] on this pin."""
    for alt_name, func in pin_info['alt_funcs'].items():
        if func and f'eMIOS_1_CH[{target_ch}]' in func:
            m = re.match(r'ALT(\d+)', alt_name)
            if m:
                return int(m.group(1))
    return None


def find_can0_alt(pin_info, direction):
    """Find which ALT mode gives CAN0_TX or CAN0_RX on this pin."""
    target = f'CAN0_{direction}'
    for alt_name, func in pin_info['alt_funcs'].items():
        if func and target in func:
            m = re.match(r'ALT(\d+)', alt_name)
            if m:
                return int(m.group(1))
    # Also check composite for IMCR-based functions (like CAN0_RX)
    if target in pin_info.get('composite', ''):
        # CAN0_RX is typically an IMCR function, check IO signal table
        return -1  # Signal it exists but ALT is in IMCR
    return None


def allocate_pins(pins):
    """Perform the complete pin allocation."""
    allocated = {}   # signal_name -> pin_info dict with assignment details
    used_pins = set()  # pin_names already assigned

    # =====================================================================
    # 1. Reserve dedicated / fixed pins
    # =====================================================================

    # JTAG (verified from datasheet)
    # Note: PTA4 (pad 170), PTC4 (pad 166), PTC5 (pad 165) may not be in the
    # GPIO pin database because the IO Signal Table has no pad number for them.
    # Reserve them regardless -- they are physically present on the package.
    jtag_pins = {
        'JTAG_TMS': ('PTA4', 170),
        'JTAG_TDO': ('PTA10', 161),
        'JTAG_TCK': ('PTC4', 166),
        'JTAG_TDI': ('PTC5', 165),
    }
    for sig, (pname, pad) in jtag_pins.items():
        if pname in pins:
            pins[pname]['assigned'] = sig
        used_pins.add(pname)
        allocated[sig] = {'pin': pname, 'pad': pad, 'func': 'JTAG', 'alt': 'dedicated'}

    # Crystal (dedicated, not GPIO)
    # EXTAL = pad 23, XTAL = pad 25 - not in GPIO table

    # CAN0 on PTC3 (pad 49, TX) and PTC2 (pad 50, RX) -- non-ADC pins
    can_pins = {
        'CAN0_TX': ('PTC3', 49, 'ALT3'),
        'CAN0_RX': ('PTC2', 50, 'IMCR'),  # Input mux, SSS=0000_0001
    }
    for sig, (pname, pad, alt) in can_pins.items():
        if pname in pins:
            pins[pname]['assigned'] = sig
            used_pins.add(pname)
            allocated[sig] = {'pin': pname, 'pad': pad, 'func': sig, 'alt': alt}

    # =====================================================================
    # 2. Safety-critical brake switches -- must be non-ADC direct GPIO
    # =====================================================================
    # PTA30 (pad 33) and PTA31 (pad 39) -- verified non-ADC
    brake_pins = {
        'BRAKE_SW1': ('PTA30', 33),
        'BRAKE_SW2': ('PTA31', 39),
    }
    for sig, (pname, pad) in brake_pins.items():
        if pname in pins:
            pins[pname]['assigned'] = sig
            used_pins.add(pname)
            allocated[sig] = {'pin': pname, 'pad': pad, 'func': 'GPIO_IN', 'alt': 'ALT0'}

    # =====================================================================
    # 3. PWM outputs -- need eMIOS1 channels on non-ADC bonded pins
    # =====================================================================
    # Strategy: find pins that have eMIOS_1 and prefer non-ADC pins.
    # Need 14 channels (12 gate driver PWM + 2 H-bridge PWM)

    pwm_signals = [
        # (signal_name, emios1_channel_preference)
        # We'll pick from available eMIOS_1 channels on non-ADC pins
        'FAN_1', 'FAN_2', 'BLOWER',
        'LOW_BEAM_L', 'LOW_BEAM_R', 'HIGH_BEAM_L', 'HIGH_BEAM_R',
        'DRL', 'INTERIOR', 'COURTESY',
        'SEAT_HEATER_L', 'SEAT_HEATER_R',
        'HBRIDGE_IN1', 'HBRIDGE_IN2',
    ]

    # Build list of candidate pins with eMIOS_1 capability, preferring non-ADC
    emios1_candidates = []
    for pname, pinfo in pins.items():
        if pname in used_pins:
            continue
        if pinfo['is_gpi']:
            continue
        for inst, ch, suffix in pinfo['emios']:
            if inst == 1:
                emios1_candidates.append((pname, ch, suffix, pinfo['has_adc'], pinfo['pad']))

    # Sort: prefer non-ADC pins, then by channel number for predictability
    emios1_candidates.sort(key=lambda x: (x[3], x[1], x[4]))

    # Remove duplicates -- each pin can provide one channel
    # Group by pin, pick the first (lowest) channel per pin
    pin_to_emios1 = {}
    for pname, ch, suffix, has_adc, pad in emios1_candidates:
        if pname not in pin_to_emios1:
            pin_to_emios1[pname] = []
        pin_to_emios1[pname].append((ch, suffix, has_adc, pad))

    # Strategy: pick non-ADC pins with unique eMIOS1 channels
    # We need 14 unique channels. Build available channel -> pin list.
    ch_to_pins = defaultdict(list)
    for pname, channels in pin_to_emios1.items():
        if pname in used_pins:
            continue
        for ch, suffix, has_adc, pad in channels:
            ch_to_pins[ch].append((pname, has_adc, pad, suffix))

    # Sort each channel's pin list: non-ADC first
    for ch in ch_to_pins:
        ch_to_pins[ch].sort(key=lambda x: (x[1], x[2]))

    # Greedily assign PWM signals to eMIOS1 channels
    # Available eMIOS1 channels (prefer non-ADC): let's find 14 non-conflicting channels
    available_channels = sorted(ch_to_pins.keys())

    # Preferred channel assignment order -- try to use channels that have
    # the most non-ADC pin options
    non_adc_ch = []
    adc_ch = []
    for ch in available_channels:
        has_non_adc = any(not has_adc for _, has_adc, _, _ in ch_to_pins[ch])
        if has_non_adc:
            non_adc_ch.append(ch)
        else:
            adc_ch.append(ch)

    # Pick 14 channels, preferring those with non-ADC pins available
    selected_channels = []
    used_ch_pins = set()

    def try_assign_channel(ch):
        """Try to assign a pin for this eMIOS1 channel. Returns (pin_name, pad) or None."""
        for pname, has_adc, pad, suffix in ch_to_pins[ch]:
            if pname not in used_pins and pname not in used_ch_pins:
                return (pname, pad, suffix)
        return None

    # First pass: assign channels that only have non-ADC options
    for ch in non_adc_ch:
        if len(selected_channels) >= 14:
            break
        result = try_assign_channel(ch)
        if result:
            pname, pad, suffix = result
            # Prefer non-ADC pin
            best = None
            for pn, ha, pa, su in ch_to_pins[ch]:
                if pn not in used_pins and pn not in used_ch_pins and not ha:
                    best = (pn, pa, su)
                    break
            if best:
                pname, pad, suffix = best
            selected_channels.append((ch, pname, pad, suffix))
            used_ch_pins.add(pname)

    # If we still need more, use ADC-capable pins
    for ch in adc_ch:
        if len(selected_channels) >= 14:
            break
        result = try_assign_channel(ch)
        if result:
            pname, pad, suffix = result
            selected_channels.append((ch, pname, pad, suffix))
            used_ch_pins.add(pname)

    # If still short, also try non_adc channels we might have missed
    if len(selected_channels) < 14:
        for ch in available_channels:
            if len(selected_channels) >= 14:
                break
            if any(c[0] == ch for c in selected_channels):
                continue
            result = try_assign_channel(ch)
            if result:
                pname, pad, suffix = result
                selected_channels.append((ch, pname, pad, suffix))
                used_ch_pins.add(pname)

    # Sort by channel number
    selected_channels.sort(key=lambda x: x[0])

    # Assign PWM signals to selected channels
    pwm_assignments = {}
    for i, sig in enumerate(pwm_signals):
        if i < len(selected_channels):
            ch, pname, pad, suffix = selected_channels[i]
            pins[pname]['assigned'] = f'PWM_{sig}'
            used_pins.add(pname)
            alt_mode = find_emios1_alt(pins[pname], ch)
            pwm_assignments[sig] = {
                'pin': pname, 'pad': pad, 'emios_ch': ch,
                'suffix': suffix, 'alt': f'ALT{alt_mode}' if alt_mode else 'ALT2',
                'func': f'eMIOS_1_CH[{ch}]_{suffix}'
            }
            allocated[f'PWM_{sig}'] = pwm_assignments[sig]

    # =====================================================================
    # 4. H-bridge control GPIO (NSLEEP output, NFAULT input)
    # =====================================================================
    # Find non-ADC GPIO pins for these two
    hbridge_gpio_sigs = [
        ('HBRIDGE_NSLEEP', 'GPIO_OUT'),
        ('HBRIDGE_NFAULT', 'GPIO_IN'),
    ]

    def find_non_adc_gpio(exclude_ports=None):
        """Find an unassigned non-ADC GPIO-capable pin."""
        candidates = []
        for pname, pinfo in pins.items():
            if pname in used_pins:
                continue
            if pinfo['is_gpi'] and 'OUT' in '':
                continue
            if pinfo['has_adc']:
                continue
            if exclude_ports and pinfo['port'] in exclude_ports:
                continue
            candidates.append((pname, pinfo['pad']))
        candidates.sort(key=lambda x: x[1])
        return candidates

    non_adc_available = find_non_adc_gpio()
    for sig, func_type in hbridge_gpio_sigs:
        for pname, pad in non_adc_available:
            if pname in used_pins:
                continue
            pinfo = pins[pname]
            if func_type == 'GPIO_OUT' and pinfo['is_gpi']:
                continue
            pins[pname]['assigned'] = sig
            used_pins.add(pname)
            allocated[sig] = {'pin': pname, 'pad': pad, 'func': func_type, 'alt': 'ALT0'}
            break

    # =====================================================================
    # 5. Switch inputs (15 total, 2 brakes already assigned)
    # =====================================================================
    switch_signals = [
        'SW_TURN_L', 'SW_TURN_R', 'SW_HIGH_BEAM', 'SW_FLASH_PASS',
        'SW_HAZARD', 'SW_HORN', 'SW_REVERSE', 'SW_AC_REQ',
        'SW_WIPER_INT', 'SW_WIPER_LO', 'SW_WIPER_HI', 'SW_WASHER',
        'SW_START_BTN',
    ]

    # Prefer non-ADC pins for switches too
    non_adc_gpio_list = []
    for pname, pinfo in pins.items():
        if pname in used_pins:
            continue
        if not pinfo['has_adc']:
            non_adc_gpio_list.append((pname, pinfo['pad'], pinfo['port']))
    non_adc_gpio_list.sort(key=lambda x: x[1])

    for sig in switch_signals:
        for pname, pad, port in non_adc_gpio_list:
            if pname in used_pins:
                continue
            pins[pname]['assigned'] = sig
            used_pins.add(pname)
            allocated[sig] = {'pin': pname, 'pad': pad, 'func': 'GPIO_IN', 'alt': 'ALT0'}
            break

    # =====================================================================
    # 6. Non-PWM gate driver outputs (34 GPIO outputs)
    # =====================================================================
    gate_signals = [
        'FUEL_PUMP', 'AC_CLUTCH',
        'TURN_L', 'TURN_R', 'BRAKE_L', 'BRAKE_R', 'REVERSE',
        'HORN', 'WIPER', 'ACCESSORY', 'FRONT_AXLE', 'LIGHT_BAR',
        'AMP_REMOTE_1', 'AMP_REMOTE_2', 'HEADUNIT_EN',
        'CAM_FRONT', 'CAM_REAR', 'CAM_SIDE', 'PARK_SENS',
        'RADAR_BSM', 'GCM_POWER', 'GPS_CELL', 'DASH_CAM',
        'FUTURE_MOD', 'ROCK_LIGHTS', 'BED_LIGHTS', 'PUDDLE_LIGHTS',
        'FUTURE_EXT',
        'EXPANSION_1', 'EXPANSION_2', 'EXPANSION_3',
        'EXPANSION_4', 'EXPANSION_5', 'EXPANSION_6',
    ]

    # Use non-ADC pins first, then ADC-capable pins
    all_gpio_out = []
    for pname, pinfo in pins.items():
        if pname in used_pins:
            continue
        if pinfo['is_gpi']:
            continue
        all_gpio_out.append((pname, pinfo['pad'], pinfo['has_adc']))
    # Sort: non-ADC first, then by pad number
    all_gpio_out.sort(key=lambda x: (x[2], x[1]))

    for sig in gate_signals:
        for pname, pad, has_adc in all_gpio_out:
            if pname in used_pins:
                continue
            pins[pname]['assigned'] = f'GATE_{sig}'
            used_pins.add(pname)
            allocated[f'GATE_{sig}'] = {'pin': pname, 'pad': pad, 'func': 'GPIO_OUT', 'alt': 'ALT0'}
            break

    # =====================================================================
    # 7. ADC inputs -- 49 channels from remaining ADC-capable pins
    # =====================================================================
    adc_signals_csense = [
        'CS_FUEL_PUMP', 'CS_FAN_1', 'CS_FAN_2', 'CS_BLOWER',
        'CS_AC_CLUTCH', 'CS_LOW_BEAM_L', 'CS_LOW_BEAM_R',
        'CS_HIGH_BEAM_L', 'CS_HIGH_BEAM_R',
        'CS_TURN_L', 'CS_TURN_R', 'CS_BRAKE_L', 'CS_BRAKE_R',
        'CS_REVERSE', 'CS_DRL', 'CS_INTERIOR', 'CS_COURTESY',
        'CS_HORN', 'CS_WIPER', 'CS_ACCESSORY', 'CS_FRONT_AXLE',
        'CS_SEAT_HEATER_L', 'CS_SEAT_HEATER_R', 'CS_LIGHT_BAR',
        # ch24 = H-bridge, uses HB_CS_ADC
        'CS_AMP_REMOTE_1', 'CS_AMP_REMOTE_2', 'CS_HEADUNIT_EN',
        'CS_CAM_FRONT', 'CS_CAM_REAR', 'CS_CAM_SIDE', 'CS_PARK_SENS',
        'CS_RADAR_BSM', 'CS_GCM_POWER', 'CS_GPS_CELL', 'CS_DASH_CAM',
        'CS_FUTURE_MOD', 'CS_ROCK_LIGHTS', 'CS_BED_LIGHTS', 'CS_PUDDLE_LIGHTS',
        'CS_FUTURE_EXT',
        'CS_EXPANSION_1', 'CS_EXPANSION_2', 'CS_EXPANSION_3',
        'CS_EXPANSION_4', 'CS_EXPANSION_5', 'CS_EXPANSION_6',
    ]  # 46 current sense

    adc_signals_sensor = [
        'BATT_ADC', '4WD_ADC', 'HB_CS_ADC',
    ]

    all_adc_signals = adc_signals_csense + adc_signals_sensor  # 49 total

    # Find all unassigned ADC-capable pins
    # Each pin may have multiple ADC channels (from different instances).
    # Build a list of (pin, pad, all_adc_options) for flexible assignment.
    adc_pin_options = []
    for pname, pinfo in pins.items():
        if pname in used_pins:
            continue
        if pinfo['has_adc']:
            # Sort ADC options: prefer precision P, then standard S, then external X
            options = []
            for inst, ch_type, ch_num, func_str in pinfo['adc']:
                priority = {'P': 0, 'S': 1, 'X': 2}.get(ch_type, 3)
                options.append((priority, inst, ch_type, ch_num, func_str))
            options.sort()
            adc_pin_options.append((pname, pinfo['pad'], options))

    # Sort by pad number for consistent allocation
    adc_pin_options.sort(key=lambda x: x[1])

    # Track used ADC (instance, channel) pairs to avoid conflicts
    used_adc_channels = set()

    adc_assignments = {}
    for sig in all_adc_signals:
        assigned = False
        for pname, pad, options in adc_pin_options:
            if pname in used_pins:
                continue
            # Try each ADC option for this pin, pick first unused channel
            for priority, adc_inst, ch_type, ch_num, func_str in options:
                adc_key = (adc_inst, ch_num)
                if adc_key not in used_adc_channels:
                    pins[pname]['assigned'] = f'ADC_{sig}'
                    used_pins.add(pname)
                    used_adc_channels.add(adc_key)
                    adc_assignments[sig] = {
                        'pin': pname, 'pad': pad,
                        'adc_inst': adc_inst, 'ch_type': ch_type,
                        'ch_num': ch_num, 'func': func_str, 'alt': 'Direct'
                    }
                    allocated[f'ADC_{sig}'] = adc_assignments[sig]
                    assigned = True
                    break
            if assigned:
                break

    return allocated, pwm_assignments, adc_assignments, pins


# =============================================================================
# Step 4: Generate C header file
# =============================================================================

def generate_header(allocated, pwm_assignments, adc_assignments, pins):
    """Generate the C header file."""
    now = datetime.now().strftime('%Y-%m-%d')

    # Count pins by function
    n_jtag = sum(1 for k in allocated if k.startswith('JTAG'))
    n_can = sum(1 for k in allocated if k.startswith('CAN'))
    n_pwm = sum(1 for k in allocated if k.startswith('PWM'))
    n_gate = sum(1 for k in allocated if k.startswith('GATE'))
    n_switch = sum(1 for k in allocated if k.startswith('SW_') or k.startswith('BRAKE'))
    n_hbridge = sum(1 for k in allocated if k.startswith('HBRIDGE'))
    n_adc = sum(1 for k in allocated if k.startswith('ADC'))
    n_total = len(allocated)

    # Build the header
    lines = []
    L = lines.append

    L('#ifndef PDCM_CONFIG_S32K358_H')
    L('#define PDCM_CONFIG_S32K358_H')
    L('')
    L('#include <stdint.h>')
    L('')
    L('// ============================================================================')
    L(f'// S32K358 HDQFP-172 Pin Allocation -- Auto-generated {now}')
    L('// ============================================================================')
    L('//')
    L('// NXP S32K358: Dual Cortex-M7 @ 240MHz, 172-pin HDQFP')
    L('// ~142 I/O pins, 5 GPIO ports (PTA-PTE), 3x SAR ADC, 3x eMIOS, 6x CAN FD')
    L('//')
    L('// SIUL2 MSCR index = port * 32 + pin_number')
    L('//   PTA0..31 = MSCR[0..31]     PTB0..31 = MSCR[32..63]')
    L('//   PTC0..31 = MSCR[64..95]    PTD0..31 = MSCR[96..127]')
    L('//   PTE0..31 = MSCR[128..159]')
    L('//')
    L(f'// Pin budget ({n_total} assigned):')
    L(f'//   {n_pwm:3d} PWM outputs       -- eMIOS1 channels -> TC4427A / DRV8876')
    L(f'//   {n_gate:3d} gate driver GPIO  -- push-pull outputs -> TC4427A')
    L(f'//   {n_hbridge:3d} H-bridge control  -- GPIO (nSLEEP, nFAULT)')
    L(f'//   {n_switch:3d} switch inputs     -- GPIO input w/ external pull-up')
    L(f'//   {n_adc:3d} ADC inputs        -- 46 current sense + 3 sensor (direct, no MUX)')
    L(f'//   {n_can:3d} CAN FD            -- CAN0 TX/RX')
    L(f'//   {n_jtag:3d} JTAG              -- TMS, TCK, TDI, TDO')
    L('//     2 crystal            -- EXTAL (pad 23) / XTAL (pad 25) [dedicated, not GPIO]')
    L('//   ~30 power/ground       -- VDD, VSS, VDDA, etc.')
    L('//')
    L('// Source: NXP S32K358 IOMUX spreadsheet (S32K358_IOMUX.xlsx)')
    L('// Generated by: s32k358_pin_allocator.py')
    L('// ============================================================================')
    L('')

    # --- PIN_PTx macros ---
    L('// --- Port/pin encoding helpers ---')
    L('// SIUL2 MSCR index = port_base + pin_number')
    L('#define PIN_PTA(n)  ((0u * 32u) + (n))')
    L('#define PIN_PTB(n)  ((1u * 32u) + (n))')
    L('#define PIN_PTC(n)  ((2u * 32u) + (n))')
    L('#define PIN_PTD(n)  ((3u * 32u) + (n))')
    L('#define PIN_PTE(n)  ((4u * 32u) + (n))')
    L('')

    # --- JTAG ---
    L('// ============================================================================')
    L('// JTAG / Debug (fixed pins -- do not reassign)')
    L('// ============================================================================')
    L('')
    L('namespace JtagPin {')
    for sig in sorted(allocated.keys()):
        if not sig.startswith('JTAG'):
            continue
        a = allocated[sig]
        pname = a['pin']
        pad = a['pad']
        m = re.match(r'PT([A-H])(\d+)', pname)
        port_l, port_n = m.group(1), m.group(2)
        short_name = sig.replace('JTAG_', '')
        L(f'    constexpr uint8_t {short_name:12s} = PIN_PT{port_l}({port_n});'
          f'    // {pname} pad {pad} -- {sig}')
    L('    // nRESET is a dedicated pin (not GPIO-muxed)')
    L('}')
    L('')

    # --- CAN ---
    L('// ============================================================================')
    L('// CAN FD 0 -- Vehicle bus (PTC3/PTC2, pads 49/50)')
    L('// ============================================================================')
    L('')
    L('namespace CanPin {')
    for sig in ['CAN0_TX', 'CAN0_RX']:
        a = allocated[sig]
        pname = a['pin']
        pad = a['pad']
        m = re.match(r'PT([A-H])(\d+)', pname)
        port_l, port_n = m.group(1), m.group(2)
        L(f'    constexpr uint8_t {sig:12s} = PIN_PT{port_l}({port_n});'
          f'    // {pname} pad {pad} -- FLEXCAN0 {a["alt"]}')
    L('}')
    L('')

    # --- PWM Outputs ---
    L('// ============================================================================')
    L('// PWM Outputs -- eMIOS1 (all 14 PWM channels on one timer instance)')
    L('// ============================================================================')
    L('// Single eMIOS instance keeps PWM timebase synchronized across all channels.')
    L('// Each pin: MSCR ALT mode selects eMIOS_1 function.')
    L('//')
    # List the PWM pin -> channel mapping in a table
    L('//  Signal            Pin     Pad   eMIOS1_CH   ALT')
    L('//  ----------------  ------  ----  ----------  ----')
    pwm_order = [
        'FAN_1', 'FAN_2', 'BLOWER',
        'LOW_BEAM_L', 'LOW_BEAM_R', 'HIGH_BEAM_L', 'HIGH_BEAM_R',
        'DRL', 'INTERIOR', 'COURTESY',
        'SEAT_HEATER_L', 'SEAT_HEATER_R',
        'HBRIDGE_IN1', 'HBRIDGE_IN2',
    ]
    for sig in pwm_order:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            L(f'//  {sig:18s}{a["pin"]:8s}{a["pad"]:5d}  CH[{a["emios_ch"]:2d}]_{a["suffix"]}   {a["alt"]}')
    L('')
    L('namespace Pin {')
    L('')
    L('    // --- PWM-capable gate driver outputs (eMIOS1) ---')
    gate_pwm = ['FAN_1', 'FAN_2', 'BLOWER',
                'LOW_BEAM_L', 'LOW_BEAM_R', 'HIGH_BEAM_L', 'HIGH_BEAM_R',
                'DRL', 'INTERIOR', 'COURTESY',
                'SEAT_HEATER_L', 'SEAT_HEATER_R']
    for sig in gate_pwm:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            pname = a['pin']
            m = re.match(r'PT([A-H])(\d+)', pname)
            port_l, port_n = m.group(1), m.group(2)
            cname = f'CH_{sig}'
            L(f'    constexpr uint8_t {cname:20s} = PIN_PT{port_l}({port_n:>2s});'
              f'  // pad {a["pad"]:3d} -- eMIOS1_CH[{a["emios_ch"]}] PWM')

    L('')
    L('    // --- Non-PWM gate driver outputs (GPIO push-pull) ---')

    # Group by tier
    tier1 = ['FUEL_PUMP', 'AC_CLUTCH', 'TURN_L', 'TURN_R', 'BRAKE_L', 'BRAKE_R',
             'REVERSE', 'HORN', 'WIPER', 'ACCESSORY', 'FRONT_AXLE', 'LIGHT_BAR']
    tier2 = ['AMP_REMOTE_1', 'AMP_REMOTE_2', 'HEADUNIT_EN']
    tier3 = ['CAM_FRONT', 'CAM_REAR', 'CAM_SIDE', 'PARK_SENS',
             'RADAR_BSM', 'GCM_POWER', 'GPS_CELL', 'DASH_CAM',
             'FUTURE_MOD', 'ROCK_LIGHTS', 'BED_LIGHTS', 'PUDDLE_LIGHTS', 'FUTURE_EXT']
    expansion = ['EXPANSION_1', 'EXPANSION_2', 'EXPANSION_3',
                 'EXPANSION_4', 'EXPANSION_5', 'EXPANSION_6']

    for tier_name, sigs in [('Tier 1', tier1), ('Tier 2 -- Enable signals', tier2),
                             ('Tier 3 -- Cameras, modules, exterior', tier3),
                             ('Expansion', expansion)]:
        L(f'    // {tier_name}')
        for sig in sigs:
            key = f'GATE_{sig}'
            if key in allocated:
                a = allocated[key]
                pname = a['pin']
                m = re.match(r'PT([A-H])(\d+)', pname)
                port_l, port_n = m.group(1), m.group(2)
                cname = f'CH_{sig}'
                L(f'    constexpr uint8_t {cname:20s} = PIN_PT{port_l}({port_n:>2s});'
                  f'  // pad {a["pad"]:3d} -- GPIO out')

    L('')
    L('    // --- H-Bridge (DRV8876) for 4WD encoder motor ---')
    for sig in ['HBRIDGE_IN1', 'HBRIDGE_IN2']:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            pname = a['pin']
            m = re.match(r'PT([A-H])(\d+)', pname)
            port_l, port_n = m.group(1), m.group(2)
            L(f'    constexpr uint8_t {sig:20s} = PIN_PT{port_l}({port_n:>2s});'
              f'  // pad {a["pad"]:3d} -- eMIOS1_CH[{a["emios_ch"]}] PWM')

    for sig in ['HBRIDGE_NSLEEP', 'HBRIDGE_NFAULT']:
        if sig in allocated:
            a = allocated[sig]
            pname = a['pin']
            m = re.match(r'PT([A-H])(\d+)', pname)
            port_l, port_n = m.group(1), m.group(2)
            direction = 'output' if 'OUT' in a['func'] else 'input'
            L(f'    constexpr uint8_t {sig:20s} = PIN_PT{port_l}({port_n:>2s});'
              f'  // pad {a["pad"]:3d} -- GPIO {direction}')

    L('')
    L('    // --- Safety-Critical Switch Inputs (direct GPIO -- never behind SPI/I2C) ---')
    for sig in ['BRAKE_SW1', 'BRAKE_SW2']:
        a = allocated[sig]
        pname = a['pin']
        m = re.match(r'PT([A-H])(\d+)', pname)
        port_l, port_n = m.group(1), m.group(2)
        label = 'Factory brake switch' if '1' in sig else 'Dedicated brake switch'
        L(f'    constexpr uint8_t BRAKE_SWITCH_{sig[-1]:1s}    = PIN_PT{port_l}({port_n:>2s});'
          f'  // pad {a["pad"]:3d} -- {label}')

    L('}')
    L('')

    # --- Switch Inputs namespace ---
    L('// ============================================================================')
    L('// Switch Inputs -- 15 direct GPIO inputs')
    L('// ============================================================================')
    L('// External pull-up + TVS + series R + RC filter on each input.')
    L('// Brake switches are duplicated in Pin:: namespace above for clarity.')
    L('')
    L('namespace SwitchPin {')
    switch_order = [
        'SW_TURN_L', 'SW_TURN_R', 'SW_HIGH_BEAM', 'SW_FLASH_PASS',
        'SW_HAZARD', 'SW_HORN', 'SW_REVERSE', 'SW_AC_REQ',
        'SW_WIPER_INT', 'SW_WIPER_LO', 'SW_WIPER_HI', 'SW_WASHER',
        'BRAKE_SW1', 'BRAKE_SW2',
        'SW_START_BTN',
    ]
    for sig in switch_order:
        if sig in allocated:
            a = allocated[sig]
            pname = a['pin']
            m = re.match(r'PT([A-H])(\d+)', pname)
            port_l, port_n = m.group(1), m.group(2)
            clean_name = sig.replace('SW_', '').replace('BRAKE_', 'BRAKE_')
            if sig.startswith('BRAKE_'):
                clean_name = sig  # keep as-is
            L(f'    constexpr uint8_t {clean_name:16s} = PIN_PT{port_l}({port_n:>2s});'
              f'  // pad {a["pad"]:3d}')
    L('')
    L('    constexpr uint8_t NUM_SWITCHES      = 15;')
    L('}')
    L('')

    # --- eMIOS Configuration ---
    L('// ============================================================================')
    L('// eMIOS Configuration -- PWM timer instance + channel mapping')
    L('// ============================================================================')
    L('// All PWM outputs on eMIOS1 instance. Same timebase = synchronized duty cycles.')
    L('')
    L('namespace EmiosCfg {')
    L('    constexpr uint8_t INSTANCE          = 1;    // eMIOS1')
    L('')
    L('    // Gate driver PWM channels (maps to eMIOS1 channel numbers)')
    for sig in gate_pwm:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            cname = f'CH_{sig}'
            L(f'    constexpr uint8_t {cname:20s} = {a["emios_ch"]:2d};'
              f'   // eMIOS1_CH[{a["emios_ch"]}]')
    L('')
    L('    // H-bridge PWM channels')
    for sig in ['HBRIDGE_IN1', 'HBRIDGE_IN2']:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            cname = f'CH_{sig}'
            L(f'    constexpr uint8_t {cname:20s} = {a["emios_ch"]:2d};'
              f'   // eMIOS1_CH[{a["emios_ch"]}]')
    L('')
    L(f'    constexpr uint8_t NUM_PWM_CHANNELS  = {len(gate_pwm)};')
    L('}')
    L('')

    # --- ADC Configuration ---
    L('// ============================================================================')
    L('// ADC Configuration -- 3x SAR ADC instances, direct (no MUX)')
    L('// ============================================================================')
    L('//')
    L('// S32K358 has 3 SAR ADC instances (ADC0, ADC1, ADC2).')
    L('// Each INA180A1 output connects directly to an MCU ADC pin -- no analog MUX.')
    L('// ADC analog inputs bypass SIUL2 digital mux (hardwired analog paths).')
    L('//')
    L('// Pin-to-ADC mapping extracted from S32K358_IOMUX.xlsx PeripheralSummaries.')
    L('// ============================================================================')
    L('')
    L('// ADC identifier: encodes instance + channel in one byte')
    L('// Upper 2 bits = ADC instance (0-2), lower 6 bits = channel number (0-63)')
    L('#define ADC_ID(instance, channel) (((instance) << 6) | ((channel) & 0x3F))')
    L('#define ADC_INSTANCE(id)          ((id) >> 6)')
    L('#define ADC_CHANNEL(id)           ((id) & 0x3F)')
    L('')

    L('namespace AdcCurrentSense {')
    L('    // Per-channel current sense ADC assignments')
    L('    // Index = OutputChannel enum value (0-46, skip 24=H-bridge)')
    L('    // Value = ADC_ID(instance, channel)')
    L('    constexpr uint8_t NUM_CHANNELS = 46;')
    L('')

    # Build the CSENSE array
    # Map signal names to output channel indices
    csense_ch_map = [
        (0,  'FUEL_PUMP'),     (1,  'FAN_1'),        (2,  'FAN_2'),
        (3,  'BLOWER'),        (4,  'AC_CLUTCH'),     (5,  'LOW_BEAM_L'),
        (6,  'LOW_BEAM_R'),    (7,  'HIGH_BEAM_L'),   (8,  'HIGH_BEAM_R'),
        (9,  'TURN_L'),        (10, 'TURN_R'),        (11, 'BRAKE_L'),
        (12, 'BRAKE_R'),       (13, 'REVERSE'),       (14, 'DRL'),
        (15, 'INTERIOR'),      (16, 'COURTESY'),      (17, 'HORN'),
        (18, 'WIPER'),         (19, 'ACCESSORY'),     (20, 'FRONT_AXLE'),
        (21, 'SEAT_HEATER_L'), (22, 'SEAT_HEATER_R'), (23, 'LIGHT_BAR'),
        (24, None),  # H-bridge -- uses HB_CS_ADC
        (25, 'AMP_REMOTE_1'),  (26, 'AMP_REMOTE_2'),  (27, 'HEADUNIT_EN'),
        (28, 'CAM_FRONT'),     (29, 'CAM_REAR'),      (30, 'CAM_SIDE'),
        (31, 'PARK_SENS'),     (32, 'RADAR_BSM'),     (33, 'GCM_POWER'),
        (34, 'GPS_CELL'),      (35, 'DASH_CAM'),      (36, 'FUTURE_MOD'),
        (37, 'ROCK_LIGHTS'),   (38, 'BED_LIGHTS'),    (39, 'PUDDLE_LIGHTS'),
        (40, 'FUTURE_EXT'),    (41, 'EXPANSION_1'),   (42, 'EXPANSION_2'),
        (43, 'EXPANSION_3'),   (44, 'EXPANSION_4'),   (45, 'EXPANSION_5'),
        (46, 'EXPANSION_6'),
    ]

    L('    constexpr uint8_t CSENSE[47] = {')
    for ch_idx, sig_name in csense_ch_map:
        if sig_name is None:
            L(f'        0,                              '
              f'// ch {ch_idx:2d}: (H-bridge -- uses HB_CS_ADC)')
            continue
        cs_key = f'CS_{sig_name}'
        if cs_key in adc_assignments:
            a = adc_assignments[cs_key]
            comma = ',' if ch_idx < 46 else ' '
            L(f'        ADC_ID({a["adc_inst"]}, {a["ch_num"]:2d}){comma}'
              f'                // ch {ch_idx:2d}: {sig_name:16s}'
              f' -- {a["pin"]} pad {a["pad"]:3d} ({a["func"]})')
        else:
            comma = ',' if ch_idx < 46 else ' '
            L(f'        0xFF{comma}                           '
              f'// ch {ch_idx:2d}: {sig_name} -- UNASSIGNED')
    L('    };')
    L('}')
    L('')

    # --- Sensor ADC ---
    L('// --- Sensor ADC ---')
    L('namespace AdcSensor {')
    sensor_map = [
        ('BATTERY',    'BATT_ADC',  'Battery voltage divider'),
        ('FOURWD_POS', '4WD_ADC',   '4WD transfer case potentiometer'),
        ('HB_CS',      'HB_CS_ADC', 'DRV8876 IPROPI current sense'),
    ]
    for cname, sig, desc in sensor_map:
        if sig in adc_assignments:
            a = adc_assignments[sig]
            L(f'    constexpr uint8_t {cname:12s} = ADC_ID({a["adc_inst"]}, {a["ch_num"]:2d});'
              f'  // {a["pin"]} pad {a["pad"]:3d} ({a["func"]}) -- {desc}')
        else:
            L(f'    constexpr uint8_t {cname:12s} = 0xFF;'
              f'  // UNASSIGNED -- {desc}')
    L('}')
    L('')

    # --- ADC Hardware Config ---
    L('// --- ADC Hardware Config ---')
    L('namespace AdcConfig {')
    L('    constexpr uint8_t  NUM_SAR_INSTANCES    = 3;     // ADC0, ADC1, ADC2')
    L('    constexpr uint16_t RESOLUTION           = 4096;  // 12-bit SAR')
    L('    constexpr uint16_t REF_MV               = 3300;  // 3.3V VDDA reference')
    L('    constexpr uint8_t  AMPLIFIER_GAIN       = 20;    // INA180A1 (gain = 20)')
    L('}')
    L('')

    # --- ADC Pin-to-Pad lookup table ---
    L('// --- ADC Pin-to-Pad Lookup (for schematic cross-reference) ---')
    L('// Maps ADC_ID -> physical pad number on HDQFP-172 package')
    L('namespace AdcPads {')
    # Group ADC assignments by instance
    for inst in range(3):
        inst_entries = []
        for sig, a in sorted(adc_assignments.items(), key=lambda x: x[1]['ch_num']):
            if a['adc_inst'] == inst:
                inst_entries.append(a)
        if inst_entries:
            L(f'    // ADC{inst}: {len(inst_entries)} channels used')
            for a in sorted(inst_entries, key=lambda x: x['ch_num']):
                L(f'    //   {a["func"]:12s} -> {a["pin"]:6s} pad {a["pad"]:3d}')
    L('}')
    L('')

    # --- Current Sense Configuration ---
    L('// ============================================================================')
    L('// Current Sense Configuration (S32K358-specific)')
    L('// ============================================================================')
    L('// No analog MUX -- all 46 INA180A1 outputs go directly to ADC pins.')
    L('// CurrentSense module reads all channels sequentially via SAR ADC scan mode.')
    L('')
    L('namespace CurrentSense {')
    L('    constexpr uint8_t  NUM_CHANNELS         = 46;    // Direct ADC (no MUX)')
    L('    constexpr uint8_t  CHANNELS_PER_MUX     = 0;     // Not applicable')
    L('    constexpr uint8_t  NUM_MUX_ICS          = 0;     // Not applicable')
    L('    constexpr uint16_t ADC_RESOLUTION       = 4096;  // 12-bit')
    L('    constexpr uint16_t ADC_REF_MV           = 3300;  // 3.3V reference')
    L('    constexpr uint8_t  AMPLIFIER_GAIN       = 20;    // INA180A1 (gain 20)')
    L('}')
    L('')

    # --- Battery Voltage Divider ---
    L('// ============================================================================')
    L('// Battery Voltage Divider')
    L('// ============================================================================')
    L('')
    L('namespace BatteryADC {')
    L('    constexpr uint16_t DIVIDER_R1_OHM       = 10000;  // R1 to +12V')
    L('    constexpr uint16_t DIVIDER_R2_OHM       = 3300;   // R2 to GND')
    L('    constexpr uint32_t DIVIDER_MULT_X1000   = 4030;   // (R1+R2)/R2 x 1000')
    L('}')
    L('')

    # --- 4WD Position ---
    L('// ============================================================================')
    L('// 4WD Position Potentiometer')
    L('// ============================================================================')
    L('')
    L('namespace FourWDPosADC {')
    L('    constexpr uint16_t POS_2HI_MV           = 300;')
    L('    constexpr uint16_t POS_A4WD_MV          = 900;')
    L('    constexpr uint16_t POS_4HI_MV           = 1500;')
    L('    constexpr uint16_t POS_NEUTRAL_MV       = 2100;')
    L('    constexpr uint16_t POS_4LO_MV           = 2700;')
    L('    constexpr uint16_t POS_TOLERANCE_MV     = 200;')
    L('}')
    L('')

    # --- Pin Allocation Summary Table ---
    L('// ============================================================================')
    L('// Pin Allocation Summary')
    L('// ============================================================================')
    L('//')

    # Build port summary
    port_usage = defaultdict(list)
    for sig, a in sorted(allocated.items()):
        pname = a['pin']
        m = re.match(r'PT([A-H])(\d+)', pname)
        if m:
            port_usage[m.group(1)].append((int(m.group(2)), sig, a))

    for port_l in ['A', 'B', 'C', 'D', 'E']:
        entries = sorted(port_usage.get(port_l, []), key=lambda x: x[0])
        if entries:
            L(f'//  PT{port_l}:')
            for pn, sig, a in entries:
                func_desc = a.get('func', 'GPIO')
                L(f'//    PT{port_l}{pn:<3d} pad {a["pad"]:3d}  {sig}  ({func_desc})')

    L('//')
    L(f'//  Total assigned: {n_total} pins')
    L(f'//    PWM outputs:     {n_pwm}')
    L(f'//    Gate driver GPIO: {n_gate}')
    L(f'//    H-bridge ctrl:   {n_hbridge}')
    L(f'//    Switch inputs:   {n_switch}')
    L(f'//    ADC inputs:      {n_adc}')
    L(f'//    CAN:             {n_can}')
    L(f'//    JTAG:            {n_jtag}')

    # Count unassigned
    unassigned = []
    for pname, pinfo in pins.items():
        if pinfo['assigned'] is None and pinfo['pad'] is not None:
            try:
                int(pinfo['pad'])
                unassigned.append(pname)
            except (ValueError, TypeError):
                pass
    L(f'//    Unassigned:      {len(unassigned)} pins')
    L('//')
    L('// ============================================================================')
    L('')
    L('#endif // PDCM_CONFIG_S32K358_H')

    return '\n'.join(lines) + '\n'


# =============================================================================
# Step 5: Generate allocation report
# =============================================================================

def generate_report(allocated, pwm_assignments, adc_assignments, pins):
    """Generate a text allocation report."""
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    lines = []
    L = lines.append

    L('=' * 80)
    L(f'  S32K358 HDQFP-172 Pin Allocation Report -- PDCM')
    L(f'  Generated: {now}')
    L(f'  Source: NXP S32K358 IOMUX spreadsheet')
    L('=' * 80)
    L('')

    # --- Summary ---
    n_jtag = sum(1 for k in allocated if k.startswith('JTAG'))
    n_can = sum(1 for k in allocated if k.startswith('CAN'))
    n_pwm = sum(1 for k in allocated if k.startswith('PWM'))
    n_gate = sum(1 for k in allocated if k.startswith('GATE'))
    n_switch = sum(1 for k in allocated if k.startswith('SW_') or k.startswith('BRAKE'))
    n_hbridge = sum(1 for k in allocated if k.startswith('HBRIDGE'))
    n_adc = sum(1 for k in allocated if k.startswith('ADC'))
    n_total = len(allocated)

    L('SUMMARY')
    L('-' * 40)
    L(f'  Total pins assigned:    {n_total}')
    L(f'  PWM outputs (eMIOS1):   {n_pwm}')
    L(f'  Gate driver GPIO out:   {n_gate}')
    L(f'  H-bridge control:       {n_hbridge}')
    L(f'  Switch inputs:          {n_switch}')
    L(f'  ADC inputs:             {n_adc}')
    L(f'  CAN FD:                 {n_can}')
    L(f'  JTAG:                   {n_jtag}')
    L('')

    # --- Dedicated pins ---
    L('DEDICATED PINS (not GPIO)')
    L('-' * 40)
    L('  EXTAL  -- pad  23  Crystal input')
    L('  XTAL   -- pad  25  Crystal output')
    L('  nRESET -- dedicated Reset')
    L('  VREFH  -- pad  16  ADC reference high')
    L('  VREFL  -- pad  17  ADC reference low')
    L('')

    # --- JTAG ---
    L('JTAG / DEBUG')
    L('-' * 40)
    for sig in sorted(allocated.keys()):
        if sig.startswith('JTAG'):
            a = allocated[sig]
            L(f'  {sig:12s} -- {a["pin"]:6s} pad {a["pad"]:3d}')
    L('')

    # --- CAN ---
    L('CAN FD')
    L('-' * 40)
    for sig in ['CAN0_TX', 'CAN0_RX']:
        a = allocated[sig]
        L(f'  {sig:12s} -- {a["pin"]:6s} pad {a["pad"]:3d}  ({a["alt"]})')
    L('')

    # --- PWM ---
    L('PWM OUTPUTS (eMIOS1)')
    L('-' * 80)
    L(f'  {"Signal":<20s} {"Pin":<8s} {"Pad":>4s}  {"eMIOS1 Ch":<12s} {"ALT":<6s} {"Suffix"}')
    L(f'  {"-"*18:<20s} {"-"*6:<8s} {"-"*4:>4s}  {"-"*10:<12s} {"-"*4:<6s} {"-"*6}')
    pwm_order = [
        'FAN_1', 'FAN_2', 'BLOWER',
        'LOW_BEAM_L', 'LOW_BEAM_R', 'HIGH_BEAM_L', 'HIGH_BEAM_R',
        'DRL', 'INTERIOR', 'COURTESY',
        'SEAT_HEATER_L', 'SEAT_HEATER_R',
        'HBRIDGE_IN1', 'HBRIDGE_IN2',
    ]
    for sig in pwm_order:
        if sig in pwm_assignments:
            a = pwm_assignments[sig]
            L(f'  {sig:<20s} {a["pin"]:<8s} {a["pad"]:4d}'
              f'  CH[{a["emios_ch"]:2d}]       {a["alt"]:<6s} {a["suffix"]}')
    L('')

    # --- Non-PWM Gate Driver Outputs ---
    L('NON-PWM GATE DRIVER OUTPUTS (GPIO)')
    L('-' * 60)
    L(f'  {"Signal":<20s} {"Pin":<8s} {"Pad":>4s}  {"ADC on pin?"}')
    L(f'  {"-"*18:<20s} {"-"*6:<8s} {"-"*4:>4s}  {"-"*10}')
    gate_order = [
        'FUEL_PUMP', 'AC_CLUTCH', 'TURN_L', 'TURN_R', 'BRAKE_L', 'BRAKE_R',
        'REVERSE', 'HORN', 'WIPER', 'ACCESSORY', 'FRONT_AXLE', 'LIGHT_BAR',
        'AMP_REMOTE_1', 'AMP_REMOTE_2', 'HEADUNIT_EN',
        'CAM_FRONT', 'CAM_REAR', 'CAM_SIDE', 'PARK_SENS',
        'RADAR_BSM', 'GCM_POWER', 'GPS_CELL', 'DASH_CAM',
        'FUTURE_MOD', 'ROCK_LIGHTS', 'BED_LIGHTS', 'PUDDLE_LIGHTS', 'FUTURE_EXT',
        'EXPANSION_1', 'EXPANSION_2', 'EXPANSION_3',
        'EXPANSION_4', 'EXPANSION_5', 'EXPANSION_6',
    ]
    for sig in gate_order:
        key = f'GATE_{sig}'
        if key in allocated:
            a = allocated[key]
            pinfo = pins.get(a['pin'], {})
            has_adc = 'YES' if pinfo.get('has_adc', False) else 'no'
            L(f'  {sig:<20s} {a["pin"]:<8s} {a["pad"]:4d}  {has_adc}')
    L('')

    # --- H-bridge control ---
    L('H-BRIDGE CONTROL')
    L('-' * 60)
    for sig in ['HBRIDGE_NSLEEP', 'HBRIDGE_NFAULT']:
        if sig in allocated:
            a = allocated[sig]
            L(f'  {sig:<20s} {a["pin"]:<8s} pad {a["pad"]:3d}  ({a["func"]})')
    L('')

    # --- Switch Inputs ---
    L('SWITCH INPUTS (GPIO)')
    L('-' * 60)
    L(f'  {"Signal":<20s} {"Pin":<8s} {"Pad":>4s}')
    L(f'  {"-"*18:<20s} {"-"*6:<8s} {"-"*4:>4s}')
    switch_order = [
        'SW_TURN_L', 'SW_TURN_R', 'SW_HIGH_BEAM', 'SW_FLASH_PASS',
        'SW_HAZARD', 'SW_HORN', 'SW_REVERSE', 'SW_AC_REQ',
        'SW_WIPER_INT', 'SW_WIPER_LO', 'SW_WIPER_HI', 'SW_WASHER',
        'BRAKE_SW1', 'BRAKE_SW2',
        'SW_START_BTN',
    ]
    for sig in switch_order:
        if sig in allocated:
            a = allocated[sig]
            L(f'  {sig:<20s} {a["pin"]:<8s} {a["pad"]:4d}')
    L('')

    # --- ADC Inputs ---
    L('ADC INPUTS (49 total: 46 current sense + 3 sensor)')
    L('-' * 80)
    L(f'  {"Signal":<20s} {"Pin":<8s} {"Pad":>4s}  {"ADC Instance":<14s} {"Channel"}')
    L(f'  {"-"*18:<20s} {"-"*6:<8s} {"-"*4:>4s}  {"-"*12:<14s} {"-"*12}')

    # Current sense
    L('')
    L('  Current Sense:')
    csense_order = [
        'CS_FUEL_PUMP', 'CS_FAN_1', 'CS_FAN_2', 'CS_BLOWER',
        'CS_AC_CLUTCH', 'CS_LOW_BEAM_L', 'CS_LOW_BEAM_R',
        'CS_HIGH_BEAM_L', 'CS_HIGH_BEAM_R',
        'CS_TURN_L', 'CS_TURN_R', 'CS_BRAKE_L', 'CS_BRAKE_R',
        'CS_REVERSE', 'CS_DRL', 'CS_INTERIOR', 'CS_COURTESY',
        'CS_HORN', 'CS_WIPER', 'CS_ACCESSORY', 'CS_FRONT_AXLE',
        'CS_SEAT_HEATER_L', 'CS_SEAT_HEATER_R', 'CS_LIGHT_BAR',
        'CS_AMP_REMOTE_1', 'CS_AMP_REMOTE_2', 'CS_HEADUNIT_EN',
        'CS_CAM_FRONT', 'CS_CAM_REAR', 'CS_CAM_SIDE', 'CS_PARK_SENS',
        'CS_RADAR_BSM', 'CS_GCM_POWER', 'CS_GPS_CELL', 'CS_DASH_CAM',
        'CS_FUTURE_MOD', 'CS_ROCK_LIGHTS', 'CS_BED_LIGHTS', 'CS_PUDDLE_LIGHTS',
        'CS_FUTURE_EXT',
        'CS_EXPANSION_1', 'CS_EXPANSION_2', 'CS_EXPANSION_3',
        'CS_EXPANSION_4', 'CS_EXPANSION_5', 'CS_EXPANSION_6',
    ]
    for sig in csense_order:
        if sig in adc_assignments:
            a = adc_assignments[sig]
            L(f'  {sig:<20s} {a["pin"]:<8s} {a["pad"]:4d}  ADC{a["adc_inst"]:<11d} {a["func"]}')

    L('')
    L('  Sensors:')
    sensor_order = ['BATT_ADC', '4WD_ADC', 'HB_CS_ADC']
    for sig in sensor_order:
        if sig in adc_assignments:
            a = adc_assignments[sig]
            L(f'  {sig:<20s} {a["pin"]:<8s} {a["pad"]:4d}  ADC{a["adc_inst"]:<11d} {a["func"]}')
    L('')

    # --- ADC distribution by instance ---
    L('ADC DISTRIBUTION BY INSTANCE')
    L('-' * 40)
    for inst in range(3):
        count = sum(1 for a in adc_assignments.values() if a['adc_inst'] == inst)
        L(f'  ADC{inst}: {count} channels')
    L('')

    # --- Unassigned pins ---
    L('UNASSIGNED PINS')
    L('-' * 60)
    unassigned = []
    for pname, pinfo in sorted(pins.items(), key=lambda x: x[1].get('pad', 999)):
        if pinfo['assigned'] is None:
            try:
                pad = int(pinfo['pad'])
                unassigned.append((pname, pad, pinfo['has_adc']))
            except (ValueError, TypeError):
                pass
    L(f'  {len(unassigned)} pins unassigned:')
    for pname, pad, has_adc in unassigned:
        adc_note = ' (ADC capable)' if has_adc else ''
        L(f'    {pname:<8s} pad {pad:3d}{adc_note}')
    L('')

    # --- Port-by-port complete listing ---
    L('COMPLETE PORT-BY-PORT ALLOCATION')
    L('=' * 80)
    for port_l in ['A', 'B', 'C', 'D', 'E']:
        L(f'\nPort PT{port_l}:')
        L(f'  {"Pin":<8s} {"Pad":>4s}  {"Assignment":<30s} {"Function":<20s} {"ADC?"}')
        L(f'  {"-"*6:<8s} {"-"*4:>4s}  {"-"*28:<30s} {"-"*18:<20s} {"-"*4}')
        port_pins = []
        for pname, pinfo in pins.items():
            m = re.match(r'PT([A-H])(\d+)', pname)
            if m and m.group(1) == port_l:
                port_pins.append((int(m.group(2)), pname, pinfo))
        port_pins.sort()
        for pn, pname, pinfo in port_pins:
            assignment = pinfo['assigned'] or '(unassigned)'
            pad = pinfo['pad']
            has_adc = 'ADC' if pinfo['has_adc'] else ''
            func = ''
            for sig, a in allocated.items():
                if a.get('pin') == pname:
                    func = a.get('func', '')
                    break
            L(f'  {pname:<8s} {pad:4d}  {assignment:<30s} {func:<20s} {has_adc}')
    L('')

    return '\n'.join(lines) + '\n'


# =============================================================================
# Main
# =============================================================================

def main():
    print("Parsing IOMUX spreadsheet...")
    pinout, gpio_pins, adc_map, emios_map = parse_iomux()
    print(f"  Found {len(pinout)} pins in Pinout sheet")
    print(f"  Found {len(gpio_pins)} GPIO entries in IO Signal Table")
    print(f"  Found {len(adc_map)} pads with ADC capability")
    print(f"  Found {len(emios_map)} pads with eMIOS capability")

    print("\nBuilding pin database...")
    pins = build_pin_db(pinout, gpio_pins, adc_map, emios_map)
    bonded = sum(1 for p in pins.values() if p['pad'] is not None)
    print(f"  {len(pins)} total GPIO pins, {bonded} bonded out in 172-pin package")

    # Count ADC-capable bonded pins
    adc_bonded = sum(1 for p in pins.values() if p['has_adc'])
    print(f"  {adc_bonded} ADC-capable bonded pins")

    # Count non-ADC bonded (excluding GPI-only)
    non_adc = sum(1 for p in pins.values() if not p['has_adc'] and not p['is_gpi'])
    print(f"  {non_adc} non-ADC GPIO pins (excl. GPI-only)")

    # Count eMIOS1-capable bonded pins
    emios1_pins = sum(1 for p in pins.values()
                      if any(inst == 1 for inst, ch, suffix in p['emios']))
    print(f"  {emios1_pins} eMIOS1-capable bonded pins")

    print("\nAllocating pins...")
    allocated, pwm_assignments, adc_assignments, pins = allocate_pins(pins)

    # Validate
    n_pwm = sum(1 for k in allocated if k.startswith('PWM'))
    n_gate = sum(1 for k in allocated if k.startswith('GATE'))
    n_switch = sum(1 for k in allocated if k.startswith('SW_') or k.startswith('BRAKE'))
    n_adc = sum(1 for k in allocated if k.startswith('ADC'))
    print(f"  PWM: {n_pwm}/14, Gate: {n_gate}/34, Switch: {n_switch}/15, ADC: {n_adc}/49")
    print(f"  Total assigned: {len(allocated)}")

    # Check for conflicts
    used_pads = {}
    for sig, a in allocated.items():
        pad = a['pad']
        if pad in used_pads:
            print(f"  ERROR: Pad {pad} conflict: {sig} vs {used_pads[pad]}")
        used_pads[pad] = sig

    print(f"\nGenerating C header: {HEADER_OUT}")
    header = generate_header(allocated, pwm_assignments, adc_assignments, pins)
    os.makedirs(os.path.dirname(HEADER_OUT), exist_ok=True)
    with open(HEADER_OUT, 'w', newline='\n') as f:
        f.write(header)
    print(f"  Written {len(header)} bytes")

    print(f"\nGenerating allocation report: {REPORT_OUT}")
    report = generate_report(allocated, pwm_assignments, adc_assignments, pins)
    os.makedirs(os.path.dirname(REPORT_OUT), exist_ok=True)
    with open(REPORT_OUT, 'w', newline='\n') as f:
        f.write(report)
    print(f"  Written {len(report)} bytes")

    print("\nDone.")


if __name__ == '__main__':
    main()
